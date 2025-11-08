"""
Strategic GTM Agent - Enhanced Version
Integrates Vertex AI (Gemini), BigQuery with intelligent data matching and write-back
"""

from flask import Flask, request, jsonify, send_file, redirect, url_for, session, render_template_string
from google.cloud import aiplatform, bigquery
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
import vertexai
from vertexai.preview.generative_models import GenerativeModel
import pandas as pd
import json
import re
import os
import threading
import uuid
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from functools import wraps
from difflib import SequenceMatcher
import uuid
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', os.urandom(24))

# OAuth Configuration
OAUTH_CLIENT_ID = os.environ.get('OAUTH_CLIENT_ID', '')
OAUTH_CLIENT_SECRET = os.environ.get('OAUTH_CLIENT_SECRET', '')

# Initialize Google Cloud clients
PROJECT_ID = os.environ.get('GCP_PROJECT_ID', '')
LOCATION = os.environ.get('GCP_LOCATION', 'us-central1')
DATASET_ID = os.environ.get('DATASET_ID', 'ca_hk_team6_ds')

# Initialize Vertex AI
if PROJECT_ID:
    try:
        vertexai.init(project=PROJECT_ID, location=LOCATION)
        bigquery_client = bigquery.Client(project=PROJECT_ID)
        print(f"✓ Initialized Vertex AI and BigQuery for project: {PROJECT_ID}")
    except Exception as e:
        print(f"⚠ Warning: Could not initialize Google Cloud services: {str(e)}")
        bigquery_client = None
else:
    print("⚠ Warning: GCP_PROJECT_ID not set")
    bigquery_client = None
batch_jobs = {}
def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_email' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': 'Authentication required', 'redirect': '/login'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def fuzzy_match_company(input_name, table_name_list):
    """
    Fuzzy match company name to find best match in table
    Returns the matched name or None
    """
    if not input_name or not table_name_list:
        return None
    
    input_name_clean = input_name.lower().strip()
    best_match = None
    best_ratio = 0.0
    
    for table_name in table_name_list:
        if not table_name:
            continue
            
        table_name_clean = str(table_name).lower().strip()
        
        # Exact match
        if input_name_clean == table_name_clean:
            return table_name
        
        # Contains match
        if input_name_clean in table_name_clean or table_name_clean in input_name_clean:
            return table_name
        
        # Fuzzy match using SequenceMatcher
        ratio = SequenceMatcher(None, input_name_clean, table_name_clean).ratio()
        if ratio > best_ratio and ratio > 0.8:  # 80% similarity threshold
            best_ratio = ratio
            best_match = table_name
    
    return best_match

def get_bigquery_context(company_name):
    """
    Retrieve relevant context from BigQuery datasets with intelligent matching
    """
    context = {
        'customer_match': None,
        'customer_data': {},
        'relevant_products': [],
        'relevant_campaigns': [],
        'relevant_sales_plays': []
    }
    
    if not bigquery_client:
        print("BigQuery client not available")
        return context
    
    try:
        # 1. Find matching customer with fuzzy matching
        query = f"""
        SELECT company_name, industry, account_manager, relationship_status, 
               last_interaction_date, auditor_firm, annual_revenue, employee_count,
               headquarters_location
        FROM `{PROJECT_ID}.{DATASET_ID}.customers`
        """
        customers_df = bigquery_client.query(query).to_dataframe()
        
        if not customers_df.empty:
            customer_names = customers_df['company_name'].tolist()
            matched_name = fuzzy_match_company(company_name, customer_names)
            
            if matched_name:
                customer_data = customers_df[customers_df['company_name'] == matched_name].iloc[0].to_dict()
                context['customer_match'] = matched_name
                context['customer_data'] = customer_data
                print(f"✓ Found customer match: {matched_name}")
                
                # Get customer industry for product/campaign matching
                customer_industry = customer_data.get('industry', '')
        
        # 2. Get relevant products based on target industries
        query = f"""
        SELECT product_name, product_category, target_industries, features, 
               competitive_advantage, base_price
        FROM `{PROJECT_ID}.{DATASET_ID}.products`
        """
        products_df = bigquery_client.query(query).to_dataframe()
        
        if not products_df.empty:
            # Filter products by industry relevance if we have customer data
            if context['customer_data']:
                customer_industry = context['customer_data'].get('industry', '').lower()
                relevant_products = products_df[
                    products_df['target_industries'].str.lower().str.contains(customer_industry, na=False, regex=False)
                ]
                if not relevant_products.empty:
                    context['relevant_products'] = relevant_products.to_dict('records')
                    print(f"✓ Found {len(context['relevant_products'])} relevant products")
            
            # If no industry match or no customer, get all products
            if not context['relevant_products']:
                context['relevant_products'] = products_df.head(5).to_dict('records')
        
        # 3. Get relevant marketing campaigns based on target industry
        query = f"""
        SELECT campaign_name, target_industry, budget_allocated, 
               conversion_rate, end_date
        FROM `{PROJECT_ID}.{DATASET_ID}.marketing_budget`
        ORDER BY conversion_rate DESC, budget_allocated DESC
        """
        campaigns_df = bigquery_client.query(query).to_dataframe()
        
        if not campaigns_df.empty:
            # Filter campaigns by industry if we have customer data
            if context['customer_data']:
                customer_industry = context['customer_data'].get('industry', '').lower()
                relevant_campaigns = campaigns_df[
                    campaigns_df['target_industry'].str.lower().str.contains(customer_industry, na=False, regex=False)
                ]
                if not relevant_campaigns.empty:
                    context['relevant_campaigns'] = relevant_campaigns.to_dict('records')
                    print(f"✓ Found {len(context['relevant_campaigns'])} relevant campaigns")
            
            # If no industry match, get top campaigns by conversion rate
            if not context['relevant_campaigns']:
                context['relevant_campaigns'] = campaigns_df.head(3).to_dict('records')
        
        # 4. Get relevant sales plays based on target industry
        query = f"""
        SELECT play_name, target_persona, target_industry, value_proposition,
               engagement_strategy, success_metrics, recommended_products
        FROM `{PROJECT_ID}.{DATASET_ID}.sales_plays`
        """
        plays_df = bigquery_client.query(query).to_dataframe()
        
        if not plays_df.empty:
            # Filter plays by industry if we have customer data
            if context['customer_data']:
                customer_industry = context['customer_data'].get('industry', '').lower()
                relevant_plays = plays_df[
                    plays_df['target_industry'].str.lower().str.contains(customer_industry, na=False, regex=False)
                ]
                if not relevant_plays.empty:
                    context['relevant_sales_plays'] = relevant_plays.to_dict('records')
                    print(f"✓ Found {len(context['relevant_sales_plays'])} relevant sales plays")
            
            # If no industry match, get all plays
            if not context['relevant_sales_plays']:
                context['relevant_sales_plays'] = plays_df.head(5).to_dict('records')
            
    except Exception as e:
        print(f"Error querying BigQuery: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return context

def create_enhanced_analysis_prompt(company, directive, bq_context):
    """
    Create comprehensive prompt with mandatory BigQuery data usage
    """
    
    # Build customer context section
    customer_section = ""
    if bq_context['customer_match']:
        customer_data = bq_context['customer_data']
        auditor_firm = customer_data.get('auditor_firm', 'Not specified')
        customer_section = f"""
**EXISTING CUSTOMER FOUND: {bq_context['customer_match']}**
- Industry: {customer_data.get('industry', 'N/A')}
- Account Manager: {customer_data.get('account_manager', 'Info Missing')}
- Relationship Status: {customer_data.get('relationship_status', 'Info Missing')}
- Last Interaction: {customer_data.get('last_interaction_date', 'Info Missing')}
- Annual Revenue: {customer_data.get('annual_revenue', 'N/A')}
- Employee Count: {customer_data.get('employee_count', 'N/A')}
- Headquarters: {customer_data.get('headquarters_location', 'N/A')}
- **Auditor Firm**: {auditor_firm}

**CRITICAL: This is an EXISTING CUSTOMER. Your analysis must reflect this relationship.**
**AUDITOR STATUS: Our records show their auditor is "{auditor_firm}". Use this in your Auditor Status field. Do NOT say "Unknown" if we have this information.**
"""
    
    # Build products section
    products_section = ""
    if bq_context['relevant_products']:
        products_section = "\n**RELEVANT PRODUCTS FROM OUR CATALOG:**\n"
        for idx, product in enumerate(bq_context['relevant_products'], 1):
            products_section += f"""
{idx}. **{product.get('product_name', 'N/A')}** ({product.get('product_category', 'N/A')})
   - Target Industries: {product.get('target_industries', 'N/A')}
   - Key Features: {product.get('features', 'N/A')}
   - Competitive Advantage: {product.get('competitive_advantage', 'N/A')}
   - Pricing Tier: {product.get('base_price', 'N/A')}
   
"""
    
    # Build campaigns section
    campaigns_section = ""
    if bq_context['relevant_campaigns']:
        campaigns_section = "\n**RELEVANT MARKETING CAMPAIGNS:**\n"
        for idx, campaign in enumerate(bq_context['relevant_campaigns'], 1):
            campaigns_section += f"""
{idx}. **{campaign.get('campaign_name', 'N/A')}**
   - Target Industry: {campaign.get('target_industry', 'N/A')}
   - Budget Allocated: {campaign.get('budget_allocated', 'N/A')}
   - Conversion Rate: {campaign.get('conversion_rate', 'N/A')}%
   - End Date: {campaign.get('end_date', 'N/A')}
   
"""
    
    # Build sales plays section
    plays_section = ""
    if bq_context['relevant_sales_plays']:
        plays_section = "\n**RELEVANT SALES PLAYS:**\n"
        for idx, play in enumerate(bq_context['relevant_sales_plays'], 1):
            plays_section += f"""
{idx}. **{play.get('play_name', 'N/A')}**
   - Target Persona: {play.get('target_persona', 'N/A')}
   - Target Industry: {play.get('target_industry', 'N/A')}
   - Value Proposition: {play.get('value_proposition', 'N/A')}
   - Engagement Strategy: {play.get('engagement_strategy', 'N/A')}
   - Success Metrics: {play.get('success_metrics', 'N/A')}
   - Recommended Products: {play.get('recommended_products', 'N/A')}
   
"""
    
    prompt = f"""You are a world-class business intelligence analyst with access to internal company data.

COMPANY TO ANALYZE: {company}
ANALYSIS DIRECTIVE: {directive}

{customer_section}

=== INTERNAL DATA FROM YOUR COMPANY SYSTEMS ===
{products_section}
{campaigns_section}
{plays_section}
===================================================

**CRITICAL INSTRUCTIONS FOR USING INTERNAL DATA:**

1. **CUSTOMER DATA PRIORITY:**
   - If this is an existing customer, emphasize the relationship throughout your analysis
   - Reference account manager, relationship status, and last interaction date in relevant sections
   - Consider annual revenue and employee count when scoring

2. **PRODUCT RECOMMENDATIONS:**
   - You MUST recommend specific products from "RELEVANT PRODUCTS" section above
   - Match products to customer needs based on target industries and features
   - Explain why each product's competitive advantage fits this customer
   - Reference pricing tier and implementation time in your recommendations. Use web search for implementation time estimate.

3. **CAMPAIGN INSIGHTS:**
   - Incorporate conversion rates from relevant campaigns into your prospect scoring
   - If a campaign targets this customer's industry, mention it as a proven approach
   - Reference budget allocation to show company commitment to this market

4. **SALES PLAY INTEGRATION:**
   - Identify which sales play(s) from above are most relevant
   - Use the value proposition from the sales play in your win themes
   - Incorporate engagement strategy from sales plays into your recommendations
   - Reference success metrics when discussing potential outcomes

5. **SCORING REQUIREMENTS:**
   - Factor in campaign conversion rates when calculating prospect score
   - Give higher scores if we have relevant products with strong competitive advantages
   - Consider relationship status if existing customer (upsell/cross-sell opportunity)
   - If high-performing campaigns exist for this industry, boost marketing readiness score

6. **EXTERNAL RESEARCH:**
   - Use your knowledge (up to April 2024) for publicly available company information
   - DO NOT make up numbers. If data unavailable, state "Unknown - needs manual research"
   - Clearly distinguish between INTERNAL DATA and PUBLIC INFORMATION

**RESPONSE FORMAT - USE THESE EXACT HEADERS:**

## 1. COMPANY OVERVIEW
- Company: [Full legal name]
- Industry: [Primary industry]
- Location: [Headquarters city, state/country]
- Employees: [Employee count or "Unknown - needs manual research"]
- Founded: [Year or "Unknown - needs manual research"]
- Status: [Public/Private or "Unknown - needs manual research"]
- Description: [2-3 sentences]
- Key Products/Services: [List 3-5]
{f"- **Relationship Status**: {bq_context['customer_data'].get('relationship_status', 'Info Missing')}" if bq_context['customer_match'] else ""}
{f"- **Account Manager**: {bq_context['customer_data'].get('account_manager', 'Info Missing')}" if bq_context['customer_match'] else ""}
{f"- **Last Interaction Date**: {bq_context['customer_data'].get('last_interaction_date', 'Info Missing')}" if bq_context['customer_match'] else ""}

## 2. FINANCIAL HEALTH
- Revenue: [Most recent annual revenue or "Unknown - needs manual research"]
- Growth Rate: [YoY growth or "Unknown - needs manual research"]
- Operating Income: [Amount or "Unknown - needs manual research"]
- Market Cap: [For public companies or "Unknown - needs manual research"]
- Cash Position: [Amount or "Unknown - needs manual research"]
- Financial Stability: [Assessment]

## 3. PROSPECT ANALYSIS

**Prospect Level:** High/Medium/Low
**Prospect Score:** [0-100 numeric value]

**Scoring Breakdown:**
- Strategic Fit: [score]/30 {f"(+5 bonus if existing customer: {'+5' if bq_context['customer_match'] else '+0'})" if bq_context['customer_match'] else ""}
- Market Readiness: [score]/25 {f"(Influenced by {bq_context['relevant_campaigns'][0]['conversion_rate']}% campaign conversion rate)" if bq_context['relevant_campaigns'] else ""}
- Financial Capacity: [score]/20
- Competitive Position: [score]/15
- Urgency/Timing: [score]/10

**Scoring Rationale:**
[Explain how internal data influenced the score, specifically mentioning:
- Campaign conversion rates if applicable
- Product-market fit based on our catalog
- Existing relationship if customer
- Sales play alignment]

**Auditor Status:** {f'✓ {bq_context["customer_data"].get("auditor_firm", "Unknown")} (from our customer records - DO NOT change this)' if bq_context['customer_match'] and bq_context["customer_data"].get("auditor_firm") else '[Based on public information or "Unknown - needs manual research"]'}

## 4. WIN THEMES

**Identify 3-5 compelling win themes based on INTERNAL DATA:**

{f"**Recommended Sales Play**: {bq_context['relevant_sales_plays'][0]['play_name']}" if bq_context['relevant_sales_plays'] else ""}
{f"**Value Proposition**: {bq_context['relevant_sales_plays'][0]['value_proposition']}" if bq_context['relevant_sales_plays'] else ""}

1. [Win Theme 1 - Must reference relevant products or campaigns]
2. [Win Theme 2 - Must reference competitive advantages from product catalog]
3. [Win Theme 3 - Must tie to successful campaign data if available]
4. [Additional themes as relevant]

## 5. RECOMMENDED SOLUTIONS

**Products from Our Catalog:**

{f"Based on the analysis, recommend products from our catalog, explaining:" if bq_context['relevant_products'] else "Recommend products based on customer needs:"}

1. **[Product Name from Catalog]**
   - **Why It Fits:** [Explain the match between the prospect's needs and this product's features and target industry.]
   - **Key Differentiators:** [Reference this product's 'competitive_advantage' from the catalog data provided.]
   - **Estimated Implementation Timeline:** [Based on the product's complexity, estimate a realistic implementation timeline (e.g., 3-6 months).]
   - **Estimated Value Proposition:** [Use the 'base_price' from the catalog and the prospect's size (employee count/revenue) to articulate the expected value or ROI.]

[Repeat for each recommended product from the catalog]

## 6. KEY PERSONNEL
[Research public information to identify key decision-makers. **Prioritize finding individuals who match the 'Target Persona' from the relevant sales play, if available.**]

- **Executive Sponsor (CEO/C-Suite):** [Name and Title]
- **Primary Decision Maker ({f"e.g., a {bq_context['relevant_sales_plays'][0]['target_persona']}" if bq_context['relevant_sales_plays'] else "e.g., VP of Operations"}):** [Name and Title]
- **Key Influencers/Department Heads:** [Name and Title]

## 7. ENGAGEMENT STRATEGY

**Based on Sales Play: {bq_context['relevant_sales_plays'][0]['play_name'] if bq_context['relevant_sales_plays'] else 'Standard Enterprise'}**

**Recommended Approach:**
{f"- Target Persona: {bq_context['relevant_sales_plays'][0]['target_persona']}" if bq_context['relevant_sales_plays'] else ""}
{f"- Engagement Strategy: {bq_context['relevant_sales_plays'][0]['engagement_strategy']}" if bq_context['relevant_sales_plays'] else ""}
{f"- Expected Product Fit: {bq_context['relevant_sales_plays'][0]['recommended_products']}" if bq_context['relevant_sales_plays'] else ""}

**Campaign Alignment:**
{f"- Align with our internal '{bq_context['relevant_campaigns'][0]['campaign_name']}' campaign where relevant." if bq_context['relevant_campaigns'] else "- No specific internal campaign to align with."}
- **Generated Key Message:** [Based on the win themes and recommended solutions, create a concise and compelling message for the target persona.]
- **Suggested Channels:** [Recommend 2-3 marketing and sales channels (e.g., LinkedIn outreach, targeted ads, industry webinar) that are suitable for delivering the key message to this prospect.]

[Additional strategic recommendations]

## 8. GO-TO-MARKET ACTION PLAN

### Immediate Actions (Week 1-2)
1. [Action tied to account manager if existing customer]
2. [Action leveraging campaign channels]
3. [Action based on engagement strategy from sales play]

### Short-term Actions (Month 1)
1. [Action referencing specific products]
2. [Action based on success metrics from sales play]
3. [Additional actions]

### Mid-term Actions (Months 2-3)
1. [Implementation-focused actions]
2. [Relationship development actions]
3. [Additional actions]

### Long-term Actions (Months 4-6)
1. [Expansion opportunities]
2. [Partnership development]
3. [Additional actions]

**Success Metrics** {f"(from {bq_context['relevant_sales_plays'][0]['play_name']} playbook):" if bq_context['relevant_sales_plays'] else ":"}
{f"- {bq_context['relevant_sales_plays'][0]['success_metrics']}" if bq_context['relevant_sales_plays'] else "- [Define specific KPIs]"}
"""
    
    return prompt

# [Previous Flask routes: /login, /auth/google, /logout remain the same]
# [Copy lines 61-241 from main_session_auth.py]

@app.route('/login')
def login():
    """Login page with Google Sign-In"""
    if 'user_email' in session:
        return redirect('/')
    
    login_html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - Strategic GTM Agent</title>
        <script src="https://accounts.google.com/gsi/client" async defer></script>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body {
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                padding: 20px;
            }
            .login-card {
                background: white;
                border-radius: 20px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                padding: 60px 40px;
                max-width: 440px;
                width: 100%;
                text-align: center;
            }
            .logo {
                width: 80px;
                height: 80px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                border-radius: 20px;
                margin: 0 auto 30px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 40px;
                color: white;
            }
            h1 {
                color: #1a202c;
                font-size: 28px;
                margin-bottom: 10px;
                font-weight: 700;
            }
            p {
                color: #718096;
                font-size: 16px;
                margin-bottom: 40px;
            }
            #buttonDiv {
                display: flex;
                justify-content: center;
                margin-bottom: 20px;
            }
            .footer {
                margin-top: 30px;
                padding-top: 30px;
                border-top: 1px solid #e2e8f0;
                color: #a0aec0;
                font-size: 14px;
            }
            .error {
                background: #fed7d7;
                color: #c53030;
                padding: 15px;
                border-radius: 10px;
                margin-bottom: 20px;
                font-size: 14px;
            }
        </style>
    </head>
    <body>
        <div class="login-card">
            <div class="logo">📊</div>
            <h1>Strategic GTM Agent</h1>
            <p>AI-Powered Prospect Analysis</p>
            
            <div id="error" class="error" style="display:none;"></div>
            
            <div id="buttonDiv"></div>
            
            <div class="footer">
                Powered by Vertex AI & BigQuery
            </div>
        </div>
        
        <script>
            function handleCredentialResponse(response) {
                fetch('/auth/google', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        credential: response.credential
                    })
                })
                .then(res => res.json())
                .then(data => {
                    if (data.success) {
                        window.location.href = '/';
                    } else {
                        document.getElementById('error').textContent = data.error || 'Authentication failed';
                        document.getElementById('error').style.display = 'block';
                    }
                })
                .catch(err => {
                    document.getElementById('error').textContent = 'Network error: ' + err.message;
                    document.getElementById('error').style.display = 'block';
                });
            }
            
            window.onload = function () {
                google.accounts.id.initialize({
                    client_id: "''' + OAUTH_CLIENT_ID + '''",
                    callback: handleCredentialResponse
                });
                
                google.accounts.id.renderButton(
                    document.getElementById("buttonDiv"),
                    { 
                        theme: "filled_blue", 
                        size: "large",
                        text: "signin_with",
                        width: 300
                    }
                );
            }
        </script>
    </body>
    </html>
    '''
    return login_html

@app.route('/auth/google', methods=['POST'])
def auth_google():
    """Handle Google Sign-In callback"""
    try:
        data = request.json
        token = data.get('credential')
        
        if not token:
            return jsonify({'success': False, 'error': 'No credential provided'}), 400
        
        idinfo = id_token.verify_oauth2_token(token, google_requests.Request(), OAUTH_CLIENT_ID)
        
        session['user_email'] = idinfo.get('email')
        session['user_name'] = idinfo.get('name')
        session['user_picture'] = idinfo.get('picture')
        
        print(f"✓ User logged in: {session['user_email']}")
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"✗ Auth error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/logout')
def logout():
    """Logout endpoint"""
    session.clear()
    return redirect('/login')

def parse_structured_data(analysis_text):
    """Parse structured data from analysis with enhanced extraction"""
    data = {
        'prospect_level': 'Medium',
        'prospect_score': 'Unknown',
        'industry': 'Unknown - needs manual research',
        'location': 'Unknown - needs manual research',
        'employees': 'Unknown - needs manual research',
        'revenue': 'Unknown - needs manual research',
        'auditor_status': 'Unknown - needs manual research',
        'win_themes': 'Unknown - needs manual research',
        'key_personnel': 'Unknown - needs manual research',
        'engagement_strategy': 'Unknown - needs manual research',
        'gtm_immediate': 'Unknown - needs manual research',
        'gtm_short_term': 'Unknown - needs manual research',
        'gtm_mid_term': 'Unknown - needs manual research',
        'gtm_long_term': 'Unknown - needs manual research',
        'recommended_solutions': 'Unknown - needs manual research'
    }
    
    # Extract from Company Overview section
    overview_match = re.search(r'## 1\. COMPANY OVERVIEW(.*?)(?=## \d+\.)', analysis_text, re.DOTALL | re.IGNORECASE)
    if overview_match:
        overview_text = overview_match.group(1)
        
        industry_match = re.search(r'Industry[:\s]*\*?\*?([^\n]+)', overview_text, re.IGNORECASE)
        if industry_match:
            data['industry'] = industry_match.group(1).strip().replace('*', '')
        
        location_match = re.search(r'Location[:\s]*\*?\*?([^\n]+)', overview_text, re.IGNORECASE)
        if location_match:
            data['location'] = location_match.group(1).strip().replace('*', '')
        
        employees_match = re.search(r'Employees[:\s]*\*?\*?([^\n]+)', overview_text, re.IGNORECASE)
        if employees_match:
            data['employees'] = employees_match.group(1).strip().replace('*', '')
    
    # Extract from Financial Health section
    financial_match = re.search(r'## 2\. FINANCIAL HEALTH(.*?)(?=## \d+\.)', analysis_text, re.DOTALL | re.IGNORECASE)
    if financial_match:
        financial_text = financial_match.group(1)
        revenue_match = re.search(r'Revenue[:\s]*\*?\*?([^\n]+)', financial_text, re.IGNORECASE)
        if revenue_match:
            data['revenue'] = revenue_match.group(1).strip().replace('*', '')
    
    # Extract from Prospect Analysis section
    prospect_match = re.search(r'## 3\. PROSPECT ANALYSIS(.*?)(?=## \d+\.)', analysis_text, re.DOTALL | re.IGNORECASE)
    if prospect_match:
        prospect_text = prospect_match.group(1)
        
        level_patterns = [
            r'\*?\*?Prospect\s+Level\*?\*?[:\s]*\*?\*?\s*([A-Za-z]+)',
            r'Prospect\s+Level[:\s]+([A-Za-z]+)',
            r'Level[:\s]*\*?\*?\s*(High|Medium|Low)',
        ]
        
        for pattern in level_patterns:
            level_match = re.search(pattern, prospect_text, re.IGNORECASE)
            if level_match:
                data['prospect_level'] = level_match.group(1).strip()
                break
        
        score_patterns = [
            r'\*?\*?Prospect\s+Score\*?\*?[:\s]*\*?\*?\s*(\d+)',
            r'Prospect\s+Score[:\s]+(\d+)',
            r'Score[:\s]*\*?\*?\s*(\d+)/100',
            r'Score[:\s]*\*?\*?\s*(\d+)\s*/\s*100',
        ]
        
        for pattern in score_patterns:
            score_match = re.search(pattern, prospect_text, re.IGNORECASE)
            if score_match:
                try:
                    data['prospect_score'] = int(score_match.group(1))
                    break
                except:
                    pass
    
    # Fallback extractions
    if data['prospect_level'] == 'Medium':
        level_patterns = [r'\*?\*?Prospect\s+Level\*?\*?[:\s]*\*?\*?\s*([A-Za-z]+)']
        for pattern in level_patterns:
            level_match = re.search(pattern, analysis_text, re.IGNORECASE)
            if level_match:
                data['prospect_level'] = level_match.group(1).strip()
                break
    
    if data['prospect_score'] == 'Unknown':
        score_patterns = [r'\*?\*?Prospect\s+Score\*?\*?[:\s]*\*?\*?\s*(\d+)']
        for pattern in score_patterns:
            score_match = re.search(pattern, analysis_text, re.IGNORECASE)
            if score_match:
                try:
                    data['prospect_score'] = int(score_match.group(1))
                    break
                except:
                    pass
    
    # Extract Auditor Status
    auditor_match = re.search(r'\*?\*?Auditor Status\*?\*?[:\s]*([^\n]+)', analysis_text, re.IGNORECASE)
    if auditor_match:
        status = auditor_match.group(1).strip()
        if 'CHECK' in status or 'DESC' in status or '⚠' in status:
            data['auditor_status'] = '⚠️ CHECK DESC'
        elif 'Other' in status or '✓' in status:
            data['auditor_status'] = '✓ Other Auditor'
        else:
            data['auditor_status'] = status
    
    # Extract sections
    win_themes_match = re.search(r'## 4\. WIN THEMES(.*?)(?=## \d+\.|$)', analysis_text, re.DOTALL | re.IGNORECASE)
    if win_themes_match:
        data['win_themes'] = win_themes_match.group(1).strip()
    
    personnel_match = re.search(r'## 6\. KEY PERSONNEL(.*?)(?=## \d+\.|$)', analysis_text, re.DOTALL | re.IGNORECASE)
    if personnel_match:
        data['key_personnel'] = personnel_match.group(1).strip()
    
    engagement_match = re.search(r'## 7\. ENGAGEMENT STRATEGY(.*?)(?=## \d+\.|$)', analysis_text, re.DOTALL | re.IGNORECASE)
    if engagement_match:
        data['engagement_strategy'] = engagement_match.group(1).strip()
    
    gtm_match = re.search(r'## 8\. GO-TO-MARKET ACTION PLAN(.*?)(?=## \d+\.|$)', analysis_text, re.DOTALL | re.IGNORECASE)
    if gtm_match:
        gtm_text = gtm_match.group(1)
        
        immediate_match = re.search(r'### Immediate Actions.*?\n(.*?)(?=###|$)', gtm_text, re.DOTALL | re.IGNORECASE)
        if immediate_match:
            data['gtm_immediate'] = immediate_match.group(1).strip()
        
        short_match = re.search(r'### Short-term Actions.*?\n(.*?)(?=###|$)', gtm_text, re.DOTALL | re.IGNORECASE)
        if short_match:
            data['gtm_short_term'] = short_match.group(1).strip()
        
        mid_match = re.search(r'### Mid-term Actions.*?\n(.*?)(?=###|$)', gtm_text, re.DOTALL | re.IGNORECASE)
        if mid_match:
            data['gtm_mid_term'] = mid_match.group(1).strip()
        
        long_match = re.search(r'### Long-term Actions.*?\n(.*?)(?=###|$)', gtm_text, re.DOTALL | re.IGNORECASE)
        if long_match:
            data['gtm_long_term'] = long_match.group(1).strip()
    
    solutions_match = re.search(r'## 5\. RECOMMENDED SOLUTIONS(.*?)(?=## \d+\.|$)', analysis_text, re.DOTALL | re.IGNORECASE)
    if solutions_match:
        data['recommended_solutions'] = solutions_match.group(1).strip()
    
    return data

def write_analysis_to_bigquery(analysis_data):
    """
    Write analysis results to analysis_complete table
    """
    if not bigquery_client:
        print("BigQuery client not available")
        return False
    
    try:
        table_id = f"{PROJECT_ID}.{DATASET_ID}.analysis_complete"
        
        # Prepare row data matching the schema
        row_data = {
            'timestamp': datetime.utcnow().isoformat(),
            'company_name': analysis_data.get('company', ''),
            'prospect_level': analysis_data.get('prospect_level', 'Unknown'),
            'prospect_score': analysis_data.get('prospect_score'),
            'industry': analysis_data.get('industry', 'Unknown'),
            'location': analysis_data.get('location', 'Unknown'),
            'employees': analysis_data.get('employees', 'Unknown'),
            'revenue': analysis_data.get('revenue', 'Unknown'),
            'auditor_status': analysis_data.get('auditor_status', 'Unknown'),
            'win_themes': analysis_data.get('win_themes', 'Unknown'),
            'key_personnel': analysis_data.get('key_personnel', 'Unknown'),
            'engagement_strategy': analysis_data.get('engagement_strategy', 'Unknown'),
            'gtm_immediate': analysis_data.get('gtm_immediate', 'Unknown'),
            'gtm_short_term': analysis_data.get('gtm_short_term', 'Unknown'),
            'gtm_mid_term': analysis_data.get('gtm_mid_term', 'Unknown'),
            'gtm_long_term': analysis_data.get('gtm_long_term', 'Unknown'),
            'recommended_solutions': analysis_data.get('recommended_solutions', 'Unknown'),
            'analysis_status': 'success',
            'full_analysis': analysis_data.get('full_analysis', ''),
            'analyzed_by': session.get('user_email', 'unknown'),
            'directive': analysis_data.get('directive', '')
        }
        
        # Insert row
        errors = bigquery_client.insert_rows_json(table_id, [row_data])
        
        if errors:
            print(f"✗ Error writing to BigQuery: {errors}")
            return False
        else:
            print(f"✓ Analysis written to BigQuery for: {analysis_data.get('company')}")
            return True
            
    except Exception as e:
        print(f"✗ Error in write_analysis_to_bigquery: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

@app.route('/')
@login_required
def home():
    """Serve the main dashboard"""
    return send_file('index.html')

@app.route('/api/user-info', methods=['GET'])
@login_required
def user_info():
    """Get current user info"""
    return jsonify({
        'success': True,
        'email': session.get('user_email'),
        'name': session.get('user_name'),
        'picture': session.get('user_picture')
    })

@app.route('/api/analyze', methods=['POST'])
@login_required
def analyze():
    """Individual company analysis with enhanced BigQuery integration"""
    try:
        if not PROJECT_ID:
            return jsonify({'success': False, 'error': 'GCP Project not configured'}), 500
        
        data = request.json
        company = data.get('company', '').strip()
        directive = data.get('directive', '').strip()
        
        if not company or not directive:
            return jsonify({'success': False, 'error': 'Company and directive required'}), 400
        
        print(f"Analyzing company: {company} (User: {session.get('user_email')})")
        
        # Get BigQuery context with fuzzy matching
        bq_context = get_bigquery_context(company)
        
        # Create enhanced prompt
        prompt = create_enhanced_analysis_prompt(company, directive, bq_context)
        
        # Call Vertex AI (Gemini)
        model = GenerativeModel('gemini-2.5-pro')
        response = model.generate_content(
            prompt,
            generation_config={
                'temperature': 0.2,
                'max_output_tokens': 8000,
            }
        )
        
        analysis_text = response.text
        
        # Parse structured data
        structured_data = parse_structured_data(analysis_text)
        
        # Add customer match info
        if bq_context['customer_match']:
            structured_data['customer_match'] = bq_context['customer_match']
            structured_data['existing_customer'] = True
            structured_data.update(bq_context['customer_data'])
        else:
            structured_data['existing_customer'] = False
        
        # Write to BigQuery analysis_complete table
        analysis_record = {
            'company': company,
            'directive': directive,
            'prospect_level': structured_data.get('prospect_level', 'Unknown'),
            'prospect_score': structured_data.get('prospect_score', 0),
            'industry': structured_data.get('industry', 'Unknown'),
            'location': structured_data.get('location', 'Unknown'),
            'employees': structured_data.get('employees', 'Unknown'),
            'revenue': structured_data.get('revenue', 'Unknown'),
            'auditor_status': structured_data.get('auditor_status', 'Unknown'),
            'win_themes': structured_data.get('win_themes', 'Unknown'),
            'key_personnel': structured_data.get('key_personnel', 'Unknown'),
            'engagement_strategy': structured_data.get('engagement_strategy', 'Unknown'),
            'gtm_immediate': structured_data.get('gtm_immediate', 'Unknown'),
            'gtm_short_term': structured_data.get('gtm_short_term', 'Unknown'),
            'gtm_mid_term': structured_data.get('gtm_mid_term', 'Unknown'),
            'gtm_long_term': structured_data.get('gtm_long_term', 'Unknown'),
            'recommended_solutions': structured_data.get('recommended_solutions', 'Unknown'),
            'full_analysis': analysis_text
        }
        write_analysis_to_bigquery(analysis_record)
        
        print(f"✓ Analysis complete: {company}")
        print(f"  Customer Match: {bq_context['customer_match'] or 'None'}")
        print(f"  Prospect Level: {structured_data['prospect_level']}")
        print(f"  Prospect Score: {structured_data['prospect_score']}")
        
        return jsonify({
            'success': True,
            'company': company,
            'directive': directive,
            'prospectLevel': structured_data['prospect_level'],
            'score': structured_data['prospect_score'],
            'auditorStatus': structured_data['auditor_status'],
            'industry': structured_data.get('industry', 'Unknown'),
            'location': structured_data.get('location', 'Unknown'),
            'employees': structured_data.get('employees', 'Unknown'),
            'analysis': analysis_text,
            'structured_data': structured_data,
            'bigquery_context': {
                'customer_match': bq_context['customer_match'],
                'products_found': len(bq_context['relevant_products']),
                'campaigns_found': len(bq_context['relevant_campaigns']),
                'sales_plays_found': len(bq_context['relevant_sales_plays'])
            },
            'source': 'vertex-ai-gemini-enhanced'
        })
        
    except Exception as e:
        print(f"✗ Error in analyze: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Analysis failed: {str(e)}'}), 500

@app.route('/api/tables/list', methods=['GET'])
@login_required
def list_tables():
    """List available BigQuery tables"""
    try:
        if not bigquery_client:
            return jsonify({'success': False, 'error': 'BigQuery not configured'}), 500
        
        dataset_ref = bigquery_client.dataset(DATASET_ID)
        tables = list(bigquery_client.list_tables(dataset_ref))
        
        table_list = [{'name': table.table_id, 'full_name': f"{DATASET_ID}.{table.table_id}"} for table in tables]
        
        return jsonify({
            'success': True,
            'tables': table_list
        })
        
    except Exception as e:
        print(f"✗ Error listing tables: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tables/<table_name>/data', methods=['POST'])
@login_required
def get_table_data(table_name):
    """Get data from specified BigQuery table with optional prospect score matching"""
    try:
        if not bigquery_client:
            return jsonify({'success': False, 'error': 'BigQuery not configured'}), 500
        
        data = request.json or {}
        company_to_match = data.get('match_company', None)
        
        # Query table data
        query = f"""
        SELECT *
        FROM `{PROJECT_ID}.{DATASET_ID}.{table_name}`
        LIMIT 1000
        """
        
        df = bigquery_client.query(query).to_dataframe()
        
        if df.empty:
            return jsonify({
                'success': True,
                'columns': [],
                'data': [],
                'row_count': 0
            })
        
        # If company matching requested, add prospect score column
        if company_to_match and 'company_name' in df.columns:
            # Get prospect scores from analysis_complete table
            score_query = f"""
            SELECT company_name, prospect_score
            FROM `{PROJECT_ID}.{DATASET_ID}.analysis_complete`
            """
            try:
                scores_df = bigquery_client.query(score_query).to_dataframe()
                
                # Add prospect_score column with fuzzy matching
                def get_prospect_score(row_company):
                    if not row_company:
                        return 'N/A'
                    matched = fuzzy_match_company(row_company, scores_df['company_name'].tolist())
                    if matched:
                        score_row = scores_df[scores_df['company_name'] == matched]
                        if not score_row.empty:
                            return score_row.iloc[0]['prospect_score']
                    return 'N/A'
                
                df['prospect_score'] = df['company_name'].apply(get_prospect_score)
                
            except Exception as e:
                print(f"Could not add prospect scores: {str(e)}")
                df['prospect_score'] = 'N/A'
        
        # Convert DataFrame to format suitable for JSON
        columns = df.columns.tolist()
        
        # Convert date columns to strings BEFORE fillna
        for col in df.columns:
            if str(df[col].dtype).startswith('dbdate') or str(df[col].dtype).startswith('dbtime'):
                # This is a db-dtypes date column - convert to string
                df[col] = df[col].astype(str)
        
        # Now safe to fill NaN with 'N/A'
        data_records = df.fillna('N/A').to_dict('records')
        
        return jsonify({
            'success': True,
            'table_name': table_name,
            'columns': columns,
            'data': data_records,
            'row_count': len(df)
        })
        
    except Exception as e:
        print(f"✗ Error getting table data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/data-explorer.html')
@login_required
def data_explorer():
    """Serve the data explorer page"""
    return send_file('data-explorer.html')
@app.route('/api/batch-analyze', methods=['POST'])
@login_required
def batch_analyze():
    """Start batch analysis of companies from uploaded file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Read the file based on extension
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type. Use CSV or Excel'}), 400
        
        # Validate required columns
        required_columns = ['company_name', 'directive']
        if not all(col in df.columns for col in required_columns):
            return jsonify({'success': False, 'error': f'File must contain columns: {", ".join(required_columns)}'}), 400
        
        # Create job ID
        job_id = str(uuid.uuid4())
        
        # Initialize job tracking
        batch_jobs[job_id] = {
            'status': 'processing',
            'total': len(df),
            'completed': 0,
            'progress': 0,
            'results': [],
            'error': None
        }
        
        # Start batch processing in background thread
        thread = threading.Thread(
            target=process_batch_analysis,
            args=(job_id, df.to_dict('records'))
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id,
            'total_companies': len(df)
        })
        
    except Exception as e:
        print(f"✗ Error in batch_analyze: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def process_batch_analysis(job_id, companies):
    """Process batch analysis in background"""
    try:
        total = len(companies)
        results = []
        
        for idx, company_data in enumerate(companies):
            try:
                company = company_data.get('company_name', '').strip()
                directive = company_data.get('directive', '').strip()
                
                if not company or not directive:
                    continue
                
                print(f"Batch analyzing {idx+1}/{total}: {company}")
                
                # Get BigQuery context
                bq_context = get_bigquery_context(company)
                
                # Create prompt
                prompt = create_enhanced_analysis_prompt(company, directive, bq_context)
                
                # Call Vertex AI
                model = GenerativeModel('gemini-2.5-pro')
                response = model.generate_content(
                    prompt,
                    generation_config={
                        'temperature': 0.2,
                        'max_output_tokens': 8000,
                    }
                )
                
                analysis_text = response.text
                
                # Parse structured data
                structured_data = parse_structured_data(analysis_text)
                
                # Add customer match info
                if bq_context['customer_match']:
                    structured_data['customer_match'] = bq_context['customer_match']
                    structured_data['existing_customer'] = True
                    structured_data.update(bq_context['customer_data'])
                else:
                    structured_data['existing_customer'] = False
                
                # Write to BigQuery
                analysis_record = {
                    'company': company,
                    'directive': directive,
                    'prospect_level': structured_data.get('prospect_level', 'Unknown'),
                    'prospect_score': structured_data.get('prospect_score', 0),
                    'industry': structured_data.get('industry', 'Unknown'),
                    'location': structured_data.get('location', 'Unknown'),
                    'employees': structured_data.get('employees', 'Unknown'),
                    'revenue': structured_data.get('revenue', 'Unknown'),
                    'auditor_status': structured_data.get('auditor_status', 'Unknown'),
                    'win_themes': structured_data.get('win_themes', 'Unknown'),
                    'key_personnel': structured_data.get('key_personnel', 'Unknown'),
                    'engagement_strategy': structured_data.get('engagement_strategy', 'Unknown'),
                    'gtm_immediate': structured_data.get('gtm_immediate', 'Unknown'),
                    'gtm_short_term': structured_data.get('gtm_short_term', 'Unknown'),
                    'gtm_mid_term': structured_data.get('gtm_mid_term', 'Unknown'),
                    'gtm_long_term': structured_data.get('gtm_long_term', 'Unknown'),
                    'recommended_solutions': structured_data.get('recommended_solutions', 'Unknown'),
                    'full_analysis': analysis_text
                }
                write_analysis_to_bigquery(analysis_record)
                
                # Add to results
                results.append({
                    'company': company,
                    'directive': directive,
                    'prospect_level': structured_data.get('prospect_level', 'Unknown'),
                    'score': structured_data.get('prospect_score', 0),
                    'analysis': analysis_text,
                    'structured_data': structured_data
                })
                
                # Update progress
                batch_jobs[job_id]['completed'] = idx + 1
                batch_jobs[job_id]['progress'] = ((idx + 1) / total) * 100
                batch_jobs[job_id]['results'] = results
                
                print(f"✓ Batch analysis complete {idx+1}/{total}: {company}")
                
            except Exception as e:
                print(f"✗ Error analyzing {company}: {str(e)}")
                continue
        
        # Sort results by score (descending)
        results.sort(key=lambda x: x.get('score', 0) if isinstance(x.get('score'), (int, float)) else 0, reverse=True)
        
        # Mark job as complete
        batch_jobs[job_id]['status'] = 'completed'
        batch_jobs[job_id]['results'] = results
        batch_jobs[job_id]['progress'] = 100
        
        print(f"✓ Batch job {job_id} completed: {len(results)} companies analyzed")
        
    except Exception as e:
        print(f"✗ Error in process_batch_analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        batch_jobs[job_id]['status'] = 'failed'
        batch_jobs[job_id]['error'] = str(e)

@app.route('/api/batch-status/<job_id>', methods=['GET'])
@login_required
def batch_status(job_id):
    """Get status of batch analysis job"""
    try:
        if job_id not in batch_jobs:
            return jsonify({'success': False, 'error': 'Job not found'}), 404
        
        job = batch_jobs[job_id]
        
        return jsonify({
            'success': True,
            'status': job['status'],
            'progress': job['progress'],
            'completed': job['completed'],
            'total': job['total'],
            'results': job['results'],
            'error': job.get('error')
        })
        
    except Exception as e:
        print(f"✗ Error in batch_status: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
@login_required
def export_excel():
    """Export batch results to Excel"""
    try:
        data = request.json
        results = data.get('results', [])
        
        if not results:
            return jsonify({'success': False, 'error': 'No results to export'}), 400
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis Results"
        
        # Define headers
        headers = [
            'Rank', 'Company', 'Prospect Level', 'Score', 'Industry', 
            'Location', 'Employees', 'Revenue', 'Auditor Status', 'Directive'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for idx, result in enumerate(results, 1):
            structured = result.get('structured_data', {})
            
            row_data = [
                idx,
                result.get('company', 'N/A'),
                result.get('prospect_level', 'N/A'),
                result.get('score', 'N/A'),
                structured.get('industry', 'N/A'),
                structured.get('location', 'N/A'),
                structured.get('employees', 'N/A'),
                structured.get('revenue', 'N/A'),
                structured.get('auditor_status', 'N/A'),
                result.get('directive', 'N/A')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx+1, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'bi_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        print(f"✗ Error in export_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
